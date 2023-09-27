import tkinter as tk
from tkinter import filedialog
from pdf2image import convert_from_path
from PIL import Image
import cv2
import numpy as np
from paddleocr import PaddleOCR
import openpyxl
import re
import locale
import tkinter.messagebox as messagebox

# Define the 'workbook' and 'sheet' variables globally
workbook = None
sheet = None

def browse_files():
    file_paths = filedialog.askopenfilenames(filetypes=[("PDF Files", "*.pdf")])
    if file_paths:
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, ", ".join(file_paths))

def insert_custom_text():
    custom_text = entry_custom_text.get()
    global sheet  # Use the global 'sheet' variable
    if sheet is None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

   # Insert custom text into the merged cells
    custom_cell = sheet.cell(row=1, column=1, value=custom_text)
    custom_cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=9)
    custom_cell.font = openpyxl.styles.Font(name='Arial', size=20, bold=True)

# Define a global variable to keep track of the image counter
image_counter = 1

def convert_pdf_to_image(pdf_file_path):
    global image_counter  # Use the global 'image_counter' variable

    images = convert_from_path(pdf_file_path, dpi=300, first_page=0, last_page=1)
    images[0].save(f'page_tnb{image_counter}.jpg', 'JPEG')
    image_counter += 1

def extract_numbering(text):
    # Define a regex pattern to match numeric characters
    pattern = r'\b\d+\b'
    # Find all occurrences of the pattern in the text
    numbers = re.findall(pattern, text)
    # Join the numbers into a single string, separated by commas
    return '.'.join(numbers)

def run_ocr_and_save_to_excel():
    global workbook, sheet  # Use the global 'workbook' and 'sheet' variables
    custom_text = entry_custom_text.get()
    pdf_file_paths = entry_file_path.get().split(", ")
    rectangles = [
        (142, 371, 268, 406),
        (1676, 327, 1912, 364),
        (1045, 626, 1280, 667),
        (1038, 451, 1220, 490),
        (1245, 451, 1422, 490),
        (898, 1400, 1279, 1580),
        (1472, 1400, 1799, 1580),
        (2001, 1400, 2358, 1580),
        (123,1311,675,1447)
    ]

    ocr = PaddleOCR(lang='en')
    all_data = []  # List to store OCR data for all files

    for pdf_file_path in pdf_file_paths:
        convert_pdf_to_image(pdf_file_path)
        image_path = f'page_tnb{image_counter-1}.jpg'
        data = []  # List to store OCR data for a single file

        for idx, rectangle in enumerate(rectangles, start=1):
            x1, y1, x2, y2 = rectangle

            # Load the image using PIL and crop the region of interest
            image = Image.open(image_path)
            cropped_image = image.crop((x1, y1, x2, y2))

            # Convert the PIL image to an OpenCV image
            open_cv_image = cv2.cvtColor(np.array(cropped_image), cv2.COLOR_RGB2BGR)

            cv2.imwrite(f'ext_imtnb{idx}.jpg', open_cv_image)

            cropped_image_path = f'ext_imtnb{idx}.jpg'

            output = ocr.ocr(cropped_image_path)[0]
            texts = [line[1][0] for line in output]
            
            # Apply extract_numbering function to specific rectangles (index 5, 6, and 7)
            if idx in (6, 7, 8):
                # Join the OCR text lines into a single string
                full_text = '\n'.join(texts)

                # Filter out only the numeric characters from the text
                numeric_text = extract_numbering(full_text)

                data.append([numeric_text])
            else:
                data.append(texts)

        all_data.append(data)  # Store OCR data for this file

    # Call the create_excel_with_header function and pass the all_data and workbook objects
    create_excel_with_header(all_data,custom_text)

    # Remove the blank sheet
    if workbook and len(workbook.sheetnames) > 0:
        blank_sheet = workbook[workbook.sheetnames[0]]
        workbook.remove(blank_sheet)
        sheet = None

def create_excel_with_header(all_data, custom_text):
    global workbook, sheet  # Use the global 'workbook' and 'sheet' variables
    if workbook is None:
        workbook = openpyxl.Workbook()
    sheet = workbook.active

    font=openpyxl.styles.Font(name='Arial',size=11)
    
    column_headers = [
        ['No.'], ['Unit'], ['No.Akaun'], ['No.Invois'], ['Tarikh Mula'], 
        ['Tarikh Akhir'],['Baki Terdahulu (RM)'], ['Caj Semasa (RM)'], ['Pelarasan (RM)'],['Jumlah Bil (RM)']
    ]

# Merge the row before the column headers from column 1 to column 10
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

 # Insert the custom text
    custom_cell = sheet.cell(row=1, column=1, value=custom_text)
    custom_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    custom_cell.font = openpyxl.styles.Font(name='Arial', size=20, bold=True,underline='single')

    # Insert column headers
    for col_idx, header in enumerate(column_headers, start=1):
        cell=sheet.cell(row=2, column=col_idx, value=header[0])
        cell.font=openpyxl.styles.Font(name='Arial',size=11,bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')  # Center-align the column header

    col3_format = openpyxl.styles.NamedStyle(name='col3_format', number_format='000000000000')
    col4_format = openpyxl.styles.NamedStyle(name='col4_format', number_format='000000000')

# Apply the column formats to the corresponding columns
    sheet.column_dimensions['C'].number_format = '000000000000'
    sheet.column_dimensions['D'].number_format = '000000000'

    # Custom currency format
    currency_format = '#,##0.00'

 # Function to apply border to a range of cells
    def apply_border(start_row, end_row, start_col, end_col):
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(border_style='thin'),
                    right=openpyxl.styles.Side(border_style='thin'),
                    top=openpyxl.styles.Side(border_style='thin'),
                    bottom=openpyxl.styles.Side(border_style='thin')
                )

    row_idx = 3
    ocr_row_count = 0  # Initialize ocr_row_count here

    total_column_7 = 0
    total_column_8 = 0
    total_column_9 = 0
    total_column_10 = 0

    # Write OCR data row by row for each file
    for file_data in all_data:
        # Insert numbering in the first column
        cell = sheet.cell(row=row_idx, column=1, value=row_idx - 2)
        
        # Check if the row has OCR data
        has_ocr_data = any(cell_data for cell_data in file_data)

        for col_idx, cell_data in enumerate(file_data, start=2):
            cell_data_str = ', '.join(cell_data)
            cell = sheet.cell(row=row_idx, column=col_idx, value=cell_data_str)
            
            if col_idx == 3:
                try:
                    numeric_value = locale.atof(cell_data_str.replace(',', ''))
                    cell.value = numeric_value
                    cell.style = col3_format
                except ValueError:
                    cell.value = cell_data_str
            elif col_idx == 4:
                try:
                    numeric_value = locale.atof(cell_data_str.replace(',', ''))
                    cell.value = numeric_value
                    cell.style = col4_format
                except ValueError:
                    cell.value = cell_data_str
            else:
                cell.font = font

            # If the cell is in column 7, 8, 9, 10 try to convert the value to float
            if col_idx in (7, 8, 9,10):
                try:
                    numeric_value = locale.atof(cell_data_str.replace(',', ''))
                    cell.value = numeric_value
                    cell.number_format = currency_format

                    # Update the total_column_7, total_column_8, total_column_9 and total_column_10
                    # only if the cell contains a valid numeric value
                    if col_idx == 7 and numeric_value is not None:
                        total_column_7 += numeric_value
                    elif col_idx == 8 and numeric_value is not None:
                        total_column_8 += numeric_value
                    elif col_idx == 9 and numeric_value is not None:
                        total_column_9 += numeric_value
                    elif col_idx == 10:
                        # Set to 0 if less than 0 or None
                        total_column_10 += max(0, numeric_value) if numeric_value is not None else 0
                except ValueError:
                    cell.value = cell_data_str
            else:
                cell.font = font

   # Apply border to all cells with OCR data (column 1 to 10), column headers, total cells, and the total number
            if has_ocr_data or row_idx == 2 or row_idx == row_idx:  # row_idx == row_idx refers to the total row
                apply_border(row_idx, row_idx, 1, 10)

  # Apply fill to the row based on whether it has OCR data or not
        if has_ocr_data:
            ocr_row_count += 1  # Increment ocr_row_count if the row has OCR data
            if ocr_row_count % 2 == 0:
                fill_color = "FFFFFF"  # No fill for even rows with OCR data
            else:
                fill_color = "E0E0E0"  # Light grey fill for odd rows with OCR data
            for col in sheet.iter_cols(min_col=1, max_col=10, min_row=row_idx, max_row=row_idx):
                for cell in col:
                    cell.fill = openpyxl.styles.PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        else:
            # No fill for rows without OCR data
            for col in sheet.iter_cols(min_col=1, max_col=10, min_row=row_idx, max_row=row_idx):
                for cell in col:
                    cell.fill = openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        row_idx += 1

# Add the word "Total" below the data in column 1 and merge across to column 6
    total_row = row_idx
    total_cell = sheet.cell(row=total_row, column=1, value="Total")
    total_cell.font = openpyxl.styles.Font(name='Arial', size=11, bold=True)
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    total_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    total_cell.fill = openpyxl.styles.PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')


    # Add the total of column 7 below the last data in column 7
    last_row_column_7 = row_idx
    cell = sheet.cell(row=last_row_column_7, column=7, value=total_column_7)
    cell.font = openpyxl.styles.Font(name='Arial', size=11, bold=True)
    cell.fill = openpyxl.styles.PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    cell.number_format = currency_format

    # Add the total of column 8 below the last data in column 8
    last_row_column_8 = row_idx
    cell = sheet.cell(row=last_row_column_8, column=8, value=total_column_8)
    cell.font = openpyxl.styles.Font(name='Arial', size=11, bold=True)
    cell.fill = openpyxl.styles.PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    cell.number_format = currency_format

    # Add the total of column 9 below the last data in column 9
    last_row_column_9 = row_idx
    cell = sheet.cell(row=last_row_column_9, column=9, value=total_column_9)
    cell.font = openpyxl.styles.Font(name='Arial', size=11, bold=True)
    cell.fill = openpyxl.styles.PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    cell.number_format = currency_format

    # Add the total of column 10 below the last data in column 10
    last_row_column_10 = row_idx
    cell = sheet.cell(row=last_row_column_10, column=10, value=total_column_10)
    cell.font = openpyxl.styles.Font(name='Arial', size=11, bold=True)
    cell.fill = openpyxl.styles.PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    cell.number_format = currency_format

# Apply border to the column headers, total cells, and last rows of columns 7, 8, 9 and 10
    apply_border(2, 2, 1, 10)  # Column headers
    apply_border(total_row, total_row, 1, 6)  # Total cell (from column 1 to 6)
    apply_border(last_row_column_7, last_row_column_7, 7, 7)  # Last row of column 7
    apply_border(last_row_column_8, last_row_column_8, 8, 8)  # Last row of column 8
    apply_border(last_row_column_9, last_row_column_9, 9, 9)  # Last row of column 9
    apply_border(last_row_column_9, last_row_column_10, 10, 10)  # Last row of column 10

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        workbook.save(file_path)
        messagebox.showinfo("Success", "Excel file has been created successfully.")
        workbook.close()  # Close the workbook after saving
        workbook = None  # Set the workbook to None after closing
    else:
        messagebox.showwarning("Warning", "File save operation was canceled.")

# Create the GUI
root = tk.Tk()
root.title("TNB OCR 2.0")

# Set the window background color to dark grey
root.config(bg='#404040')

# Label font and text color
label_font = ("Times New Roman", 12)
label_text_color = "white"

# Label background color
label_bg_color = '#404040'

# Label for the file paths entry
label_file_path = tk.Label(root, text="PDF Files Paths:", font=label_font, bg=label_bg_color, fg=label_text_color)
label_file_path.grid(row=0, column=0, padx=5, pady=5, sticky="w")

# Entry widget to show the selected file paths
entry_file_path = tk.Entry(root, width=50, font=label_font, highlightthickness=1, bd=0)
entry_file_path.grid(row=0, column=1, padx=5, pady=5, sticky="e")

# Browse button to select PDF files
button_browse_files = tk.Button(root, text="Browse PDFs", command=browse_files, font=("Times New Roman", 12, "bold"), bg="lightblue", fg="black", activebackground="blue", activeforeground="white")
button_browse_files.grid(row=0, column=2, padx=5, pady=5)

# Label for the custom text entry
label_custom_text = tk.Label(root, text="Title:[e.g.: BEV TNB JULY 2023 (VACANT UNIT)]", font=label_font, bg=label_bg_color, fg=label_text_color)
label_custom_text.grid(row=1, column=0, padx=5, pady=5, sticky="w")

# Entry widget to allow user input for custom text
entry_custom_text = tk.Entry(root, width=50, font=label_font, highlightthickness=1, bd=0)
entry_custom_text.grid(row=1, column=1, padx=5, pady=5, sticky="e")

# Label for the sheet name entry
label_sheet_name = tk.Label(root, text="Sheet Name:", font=label_font, bg=label_bg_color, fg=label_text_color)
label_sheet_name.grid(row=2, column=0, padx=5, pady=5, sticky="w")

# Entry widget to allow user input for the sheet name
entry_sheet_name = tk.Entry(root, width=50, font=label_font, highlightthickness=1, bd=0)
entry_sheet_name.grid(row=2, column=1, padx=5, pady=5, sticky="e")

# OCR and Save to Excel button
button_ocr_and_save_to_excel = tk.Button(root, text="OCR and Save to Excel", command=run_ocr_and_save_to_excel, font=("Times New Roman", 14, "bold"), bg="green", fg="white", activebackground="darkgreen", activeforeground="white")
button_ocr_and_save_to_excel.grid(row=3, column=0, columnspan=3, padx=5, pady=5)

# Set the window size to fit the content
root.geometry("")

root.mainloop()